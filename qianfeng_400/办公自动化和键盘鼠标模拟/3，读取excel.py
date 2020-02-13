#xls
# openpyxl 处理xlsx格式，不能处理xls文件

from openpyxl.reader.excel import load_workbook


def readXlsxFile(path):
    # 得到整张表对象
    file = load_workbook(filename=path)
    # 得到所有sheet名称
    sheets = file.sheetnames
    # 得到一个sheet对象
    sheet0 = file[sheets[0]]
    # 得到行数、列数、表明
    # print(sheet0.max_row)
    # print(sheet0.max_column)
    print(sheet0.title)
    # 读取一张表中的数据
    for rowNum in range(1,sheet0.max_row+1):
        rowList = []
        for columnNum in range(1,sheet0.max_column+1):
            value = sheet0.cell(row = rowNum,column = columnNum).value
            rowList.append(value)
        print(rowList)


path = r"GeoLite2-ASN-Blocks-IPv4.xlsx"
readXlsxFile(path)