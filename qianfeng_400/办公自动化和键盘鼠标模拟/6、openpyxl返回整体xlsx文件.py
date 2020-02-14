
from openpyxl.reader.excel import load_workbook


def readXlsxFile(path):
    # 得到整张表对象
    file = load_workbook(filename=path)
    # 得到所有sheet名称
    sheets = file.sheetnames
    # 整张表数据存在xlsxInfo中
    # 一张表的数据sheetInfo中
    # 一行数据存在rowInfo中

    # 这个声明不能写在下面这个循环里面，因为那样的话，每次循环就会把前面的内容清空了，再添加新内容
    xlsxInfo = {}
    for i in range(len(sheets)):
        sheet = file[sheets[i]]
        sheetInfo = []
        for rowNum in range(1,sheet.max_row+1):
            rowInfo = [] # 临时用来存放一张行数据，所以要求，一行添加完成之后就要把前一行的数据删除，不能永久保存，下载
            for columnNum in range(1,sheet.max_column+1):
                rowInfo.append(sheet.cell(row = rowNum,column = columnNum).value)
            sheetInfo.append(rowInfo)
        xlsxInfo[sheet] = sheetInfo
    print(xlsxInfo)


path = r"aa13.xlsx"
print(readXlsxFile(path))